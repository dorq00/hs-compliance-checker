"""
PPI Quota Dashboard — ppi_dashboard.py
=======================================
Standalone script. Do not import from run.py.

Logic:
  - Category == 'SVC' in approval file → CVC bank quota
  - Category != 'SVC' (REF, TV, WM, AC ...) → HSBC bank quota
  - Consumed qty pulled from processed output files (PPI_CHECK sheet, Bank column)

Run:
    python ppi_dashboard.py

Output:  THO_Checker/output/PPI_QUOTA_DASHBOARD_YYYY-MM.xlsx
"""

import sys
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Run:  pip install pandas openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────

ROOT          = Path(__file__).parent
OUTPUT_DIR    = ROOT / "output"
APPROVAL_FILE = ROOT / "ppi" / "ppi_approval_q1_2026.xlsx"
PPI_CVC_FILE  = ROOT / "ppi" / "ppi_cvc.xlsx"
PPI_HSBC_FILE = ROOT / "ppi" / "ppi_hsbc.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

THIN        = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))
HDR_FILL    = _fill("2E4057")
GREEN_FILL  = _fill("C6EFCE")
RED_FILL    = _fill("FFC7CE")
ORANGE_FILL = _fill("FFEB9C")
GREY_FILL   = _fill("F2F2F2")
GREEN_FONT  = Font(color="006100", bold=True)
RED_FONT    = Font(color="9C0006", bold=True)
ORANGE_FONT = Font(color="9C5700", bold=True)

# ── HS normalization (mirrors run.py) ─────────────────────────────────────────

def normalize_hs(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "-"
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return "-"
    s = s.replace(".", "").replace(" ", "")
    if s.isdigit():
        if len(s) == 11 and s.endswith("0"):
            s = s[:10]
        return s.zfill(10)
    return s

# ── Load quotas from approval file ───────────────────────────────────────────

def load_quota(path: Path):
    """
    Read PM Data (2), filter Remark == 'Approved'.
    SVC category  → CVC quota
    Non-SVC       → HSBC quota
    Returns two dicts: {hs_code: {qty, description}} for CVC and HSBC.
    """
    if not path.exists():
        print(f"ERROR: Approval file not found:\n  {path}")
        sys.exit(1)

    df = pd.read_excel(path, sheet_name="PM Data (2)", header=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").fillna("")

    def _col(*keywords):
        for c in df.columns:
            if any(k in c.lower().replace("\n", " ") for k in keywords):
                return c
        return None

    cat_col  = _col("category")
    hs_col   = _col("sous-position", "sous position")
    desc_col = _col("désignation", "designation")
    qty_col  = _col("quantité à importer", "quantite a importer")
    rmk_col  = _col("remark")

    if not hs_col or not qty_col:
        print("ERROR: Could not locate HS or Qty columns in PM Data (2).")
        sys.exit(1)

    # Filter Approved
    if rmk_col:
        df = df[df[rmk_col].str.strip().str.lower() == "approved"].reset_index(drop=True)
    else:
        print("  WARNING: Remark column not found — no filter applied.")

    df["_HS"]   = df[hs_col].apply(normalize_hs)
    df["_Desc"] = df[desc_col].str.strip() if desc_col else ""
    df["_Qty"]  = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    df["_Cat"]  = df[cat_col].str.strip().str.upper() if cat_col else ""

    df = df[df["_HS"] != "-"]

    cvc_dict  = {}
    hsbc_dict = {}

    for _, row in df.iterrows():
        hs   = row["_HS"]
        qty  = row["_Qty"]
        desc = row["_Desc"]
        cat  = row["_Cat"]

        if cat == "SVC":
            if hs in cvc_dict:
                cvc_dict[hs]["qty"] += qty
            else:
                cvc_dict[hs] = {"qty": qty, "desc": desc}
        else:
            if hs in hsbc_dict:
                hsbc_dict[hs]["qty"] += qty
            else:
                hsbc_dict[hs] = {"qty": qty, "desc": desc}

    print(f"  Quota CVC  : {len(cvc_dict)} HS codes (SVC rows)")
    print(f"  Quota HSBC : {len(hsbc_dict)} HS codes (non-SVC rows)")
    return cvc_dict, hsbc_dict

# ── Scan output files for consumed qty ───────────────────────────────────────

def load_consumption(output_dir: Path) -> pd.DataFrame:
    """
    Scan THO_CHECK_*.xlsx in output/YYYY-MM/.
    PPI_CHECK sheet: sum Qty per HS code per Bank where PPI Status == IN PPI.
    Returns DataFrame: HS Code | Consumed CVC | Consumed HSBC
    """
    records = []
    xlsx_files = sorted(output_dir.rglob("THO_CHECK_*.xlsx"))

    if not xlsx_files:
        print("  Consumption: no output files found — consumed qtys will be 0.")
        return pd.DataFrame(columns=["HS Code", "Consumed CVC", "Consumed HSBC"])

    for fpath in xlsx_files:
        try:
            xf = pd.ExcelFile(fpath, engine="openpyxl")
            if "PPI_CHECK" not in xf.sheet_names:
                continue
            df = xf.parse("PPI_CHECK", dtype=str).fillna("")
            df.columns = [str(c).strip() for c in df.columns]
            if "PPI Status" not in df.columns or "HS Invoice" not in df.columns:
                continue

            df_in = df[df["PPI Status"] == "IN PPI"].copy()
            if df_in.empty:
                continue

            df_in["_HS"]   = df_in["HS Invoice"].apply(normalize_hs)
            df_in["_Bank"] = df_in["Bank"].str.strip().str.upper() if "Bank" in df_in.columns else "UNKNOWN"
            df_in["_Qty"]  = pd.to_numeric(df_in["Qty"], errors="coerce").fillna(0) if "Qty" in df_in.columns else 0

            for _, row in df_in.iterrows():
                records.append({"HS Code": row["_HS"], "Bank": row["_Bank"], "Qty": row["_Qty"]})
        except Exception as e:
            print(f"  WARNING: could not read {fpath.name} — {e}")

    print(f"  Scanned  : {len(xlsx_files)} output file(s)")

    if not records:
        return pd.DataFrame(columns=["HS Code", "Consumed CVC", "Consumed HSBC"])

    df_all = pd.DataFrame(records)
    pivot  = (
        df_all.groupby(["HS Code", "Bank"])["Qty"]
        .sum().unstack(fill_value=0).reset_index()
    )
    pivot.columns.name = None
    for col in ("CVC", "HSBC"):
        if col not in pivot.columns:
            pivot[col] = 0
    return pivot.rename(columns={"CVC": "Consumed CVC", "HSBC": "Consumed HSBC"})[
        ["HS Code", "Consumed CVC", "Consumed HSBC"]
    ]

# ── Build dashboard DataFrame ─────────────────────────────────────────────────

def build_dashboard(cvc_dict: dict, hsbc_dict: dict, consumption: pd.DataFrame) -> pd.DataFrame:
    # Union of all HS codes across both quota dicts
    all_hs = sorted(set(cvc_dict) | set(hsbc_dict))

    rows = []
    for hs in all_hs:
        cvc_entry  = cvc_dict.get(hs,  {"qty": 0, "desc": ""})
        hsbc_entry = hsbc_dict.get(hs, {"qty": 0, "desc": ""})
        desc = cvc_entry["desc"] or hsbc_entry["desc"]
        rows.append({
            "HS Code":       hs,
            "Description":   desc,
            "Q1 Auth CVC":   cvc_entry["qty"],
            "Q1 Auth HSBC":  hsbc_entry["qty"],
        })

    df = pd.DataFrame(rows)
    df = df.merge(consumption, on="HS Code", how="left").fillna(0)

    for col in ("Q1 Auth CVC", "Q1 Auth HSBC", "Consumed CVC", "Consumed HSBC"):
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["Remaining CVC"]  = (df["Q1 Auth CVC"]  - df["Consumed CVC"]).clip(lower=0)
    df["Remaining HSBC"] = (df["Q1 Auth HSBC"] - df["Consumed HSBC"]).clip(lower=0)

    def pct(consumed, auth):
        return round(consumed / auth * 100, 1) if auth > 0 else 0.0

    df["% CVC Used"]  = df.apply(lambda r: pct(r["Consumed CVC"],  r["Q1 Auth CVC"]),  axis=1)
    df["% HSBC Used"] = df.apply(lambda r: pct(r["Consumed HSBC"], r["Q1 Auth HSBC"]), axis=1)

    return df[["HS Code", "Description",
               "Q1 Auth CVC",  "Consumed CVC",  "Remaining CVC",  "% CVC Used",
               "Q1 Auth HSBC", "Consumed HSBC", "Remaining HSBC", "% HSBC Used"]]

# ── Write Excel ───────────────────────────────────────────────────────────────

def _pct_style(ws, ri, ci, pct_val):
    cell = ws.cell(ri, ci)
    if pct_val >= 90:
        cell.fill = RED_FILL;    cell.font = RED_FONT
    elif pct_val >= 70:
        cell.fill = ORANGE_FILL; cell.font = ORANGE_FONT
    else:
        cell.fill = GREEN_FILL;  cell.font = GREEN_FONT

def _rem_style(ws, ri, ci, remaining, auth):
    if auth > 0 and remaining == 0:
        ws.cell(ri, ci).fill = RED_FILL;    ws.cell(ri, ci).font = RED_FONT
    elif auth > 0 and remaining < auth * 0.1:
        ws.cell(ri, ci).fill = ORANGE_FILL; ws.cell(ri, ci).font = ORANGE_FONT

def write_excel(df: pd.DataFrame, out_path: Path, run_time: str):
    wb = openpyxl.Workbook()

    # ── DASHBOARD summary sheet ───────────────────────────────────
    ws_d = wb.active
    ws_d.title = "DASHBOARD"
    ws_d.sheet_properties.tabColor = "2E4057"
    ws_d.sheet_view.showGridLines = False

    ws_d.merge_cells("A1:D1")
    tc = ws_d["A1"]
    tc.value = "PPI QUOTA DASHBOARD — Q1 2026"
    tc.font  = Font(bold=True, size=14, color="FFFFFF")
    tc.fill  = _fill("2E4057")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_d.row_dimensions[1].height = 32

    for i, (k, v) in enumerate([("Generated", run_time), ("Scope", "Q1 2026")], 3):
        ws_d.cell(i, 1, k).font = Font(bold=True)
        ws_d.cell(i, 2, v)

    summary = [
        ("",                     "CVC",                                   "HSBC"),
        ("HS Codes",             int((df["Q1 Auth CVC"] > 0).sum()),      int((df["Q1 Auth HSBC"] > 0).sum())),
        ("Q1 Auth Qty",          int(df["Q1 Auth CVC"].sum()),             int(df["Q1 Auth HSBC"].sum())),
        ("Consumed",             int(df["Consumed CVC"].sum()),            int(df["Consumed HSBC"].sum())),
        ("Remaining",            int(df["Remaining CVC"].sum()),           int(df["Remaining HSBC"].sum())),
    ]
    for ci, h in enumerate(["Metric", "CVC", "HSBC"], 1):
        c = ws_d.cell(6, ci, h)
        c.fill = _fill("2E4057"); c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center"); c.border = THIN
    for ri, row_data in enumerate(summary[1:], 7):
        for ci, val in enumerate(row_data, 1):
            c = ws_d.cell(ri, ci, val)
            c.border = THIN
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center")

    ws_d.column_dimensions["A"].width = 18
    ws_d.column_dimensions["B"].width = 14
    ws_d.column_dimensions["C"].width = 14

    # ── QUOTA_DETAIL sheet ────────────────────────────────────────
    ws = wb.create_sheet("QUOTA_DETAIL")
    ws.sheet_properties.tabColor = "4472C4"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = list(df.columns)
    col_widths = {
        "HS Code": 14, "Description": 40,
        "Q1 Auth CVC": 13, "Consumed CVC": 13, "Remaining CVC": 13, "% CVC Used": 11,
        "Q1 Auth HSBC": 13, "Consumed HSBC": 13, "Remaining HSBC": 13, "% HSBC Used": 11,
    }

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = HDR_FILL; c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 13)
    ws.row_dimensions[1].height = 28

    pct_cvc_ci  = headers.index("% CVC Used")  + 1
    pct_hsbc_ci = headers.index("% HSBC Used") + 1
    rem_cvc_ci  = headers.index("Remaining CVC")  + 1
    rem_hsbc_ci = headers.index("Remaining HSBC") + 1

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.border = THIN
            c.alignment = Alignment(vertical="center", horizontal="left" if ci == 2 else "center")
            if ri % 2 == 0:
                c.fill = GREY_FILL

        _pct_style(ws, ri, pct_cvc_ci,  row["% CVC Used"])
        _pct_style(ws, ri, pct_hsbc_ci, row["% HSBC Used"])
        _rem_style(ws, ri, rem_cvc_ci,  row["Remaining CVC"],  row["Q1 Auth CVC"])
        _rem_style(ws, ri, rem_hsbc_ci, row["Remaining HSBC"], row["Q1 Auth HSBC"])

    wb.save(out_path)

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  PPI QUOTA DASHBOARD")
    print("=" * 60)

    run_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    print("\n  Loading quotas from approval file...")
    cvc_dict, hsbc_dict = load_quota(APPROVAL_FILE)

    print("\n  Scanning output files for consumption...")
    consumption = load_consumption(OUTPUT_DIR)

    print("\n  Building dashboard...")
    df = build_dashboard(cvc_dict, hsbc_dict, consumption)

    out_name = f"PPI_QUOTA_DASHBOARD_{datetime.now().strftime('%Y-%m')}.xlsx"
    out_path = OUTPUT_DIR / out_name
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    write_excel(df, out_path, run_time)
    print(f"\n  [DONE]  {out_path}")

    try:
        import os
        os.startfile(out_path)
    except Exception:
        pass


if __name__ == "__main__":
    main()
