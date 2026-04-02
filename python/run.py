"""
THO_Checker — run.py
====================
Drop your invoice in  invoices/incoming/
Double-click          START_CHECKER.bat
Answer CVC or HSBC
Done — output lands in output/YYYY-MM/

FOLDER STRUCTURE:
  THO_Checker/
  ├── run.py                   ← this file
  ├── START_CHECKER.bat        ← double-click launcher
  ├── db/
  │   └── db.xlsx              ← swap anytime, keep the same filename
  ├── ppi/
  │   ├── ppi_cvc.xlsx
  │   └── ppi_hsbc.xlsx
  ├── invoices/
  │   └── incoming/            ← drop invoice here before running
  └── output/
      └── 2026-03/             ← auto-created, organized by month
"""

import os
import sys
from datetime import datetime
from pathlib import Path

# ── Dependencies ──────────────────────────────────────────────────────────────

try:
    import pandas as pd
    import numpy as np
except ImportError:
    print("ERROR: pandas not installed.  Run:  pip install pandas openpyxl")
    input("\nPress Enter to close...")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  Run:  pip install openpyxl")
    input("\nPress Enter to close...")
    sys.exit(1)

# ── Paths — relative to this script's location ───────────────────────────────

ROOT       = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
DB_FILE    = ROOT / "db"    / "db.xlsx"
PPI_CVC    = ROOT / "ppi"   / "ppi_cvc.xlsx"
PPI_HSBC   = ROOT / "ppi"   / "ppi_hsbc.xlsx"
INCOMING   = ROOT / "invoices" / "incoming"
OUTPUT_DIR = ROOT / "output"

# ── Colors ────────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

GREEN_FILL  = _fill("C6EFCE")
RED_FILL    = _fill("FFC7CE")
ORANGE_FILL = _fill("FFEB9C")
GREY_FILL   = _fill("F2F2F2")

GREEN_FONT  = Font(color="006100", bold=True)
RED_FONT    = Font(color="9C0006", bold=True)
ORANGE_FONT = Font(color="9C5700", bold=True)

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

VERDICT_STYLE = {
    "MATCH":               (GREEN_FILL,  GREEN_FONT),
    "IN PPI":              (GREEN_FILL,  GREEN_FONT),
    "MISMATCH":            (RED_FILL,    RED_FONT),
    "NOT IN PPI":          (RED_FILL,    RED_FONT),
    "NEW ITEM":            (ORANGE_FILL, ORANGE_FONT),
    "CHECK - HS MISMATCH": (RED_FILL,    RED_FONT),
}

TAB_COLORS = {
    "DASHBOARD":  "70AD47",
    "HS_CHECK":   "375623",
    "PPI_CHECK":  "7030A0",
    "THO_OUTPUT": "C55A11",
    "PPI_REF":    "4472C4",
}

HEADER_FILLS = {
    "DASHBOARD":  _fill("2E4057"),
    "HS_CHECK":   _fill("375623"),
    "PPI_CHECK":  _fill("7030A0"),
    "THO_OUTPUT": _fill("C55A11"),
    "PPI_REF":    _fill("4472C4"),
}

# ── HS normalization ──────────────────────────────────────────────────────────

def normalize_hs(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "-"
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return "-"
    s = s.replace(".", "").replace(" ", "")
    if s.isdigit():
        if len(s) == 11 and s.endswith("0"):
            s = s[:10]
        return s.zfill(10)
    return s

# ── Excel reader ──────────────────────────────────────────────────────────────

def _read_excel(path, required_cols):
    """Try header row 0 then 1. Falls back to headerless assignment."""
    for hr in (0, 1):
        df = pd.read_excel(path, header=hr, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all").fillna("")
        df = df.loc[:, ~df.columns.str.match(r'^(nan|Unnamed:)')]
        if all(any(req.lower() in c.lower() for c in df.columns) for req in required_cols):
            return df
    raise ValueError(
        f"Could not find columns {required_cols} in {Path(path).name}\n"
        f"  Found: {list(df.columns)}"
    )

# ── Loaders ───────────────────────────────────────────────────────────────────

def load_invoice(path):
    try:
        df = _read_excel(path, ["Part No", "HS Code"])
        df.columns = [c.replace(" (Invoice)", "") for c in df.columns]
    except ValueError:
        df = pd.read_excel(path, header=None, dtype=str).fillna("")
        names = ["Line N°", "Part No", "Designation", "Class", "HS Code"]
        df.columns = names[:len(df.columns)]
    # Detect and normalize any Qty column — optional, non-breaking
    qty_col = next(
        (c for c in df.columns if any(k in c.lower() for k in ("qty", "quant", "qté", "qte")) and c != "Qty"),
        None,
    )
    if qty_col:
        df = df.rename(columns={qty_col: "Qty"})
    print(f"  Invoice  : {len(df)} rows")
    return df


def load_db(path):
    try:
        df = _read_excel(path, ["Part No", "HS Code"])
    except ValueError:
        df = pd.read_excel(path, header=None, dtype=str).fillna("")
        names = ["Part No", "HS Code", "Designation", "Class"]
        df.columns = names[:len(df.columns)]
    df["HS Code"] = df["HS Code"].apply(normalize_hs)
    lookup = {
        str(r["Part No"]).strip(): r["HS Code"]
        for _, r in df.iterrows()
        if str(r["Part No"]).strip() not in ("", "nan", "none")
    }
    print(f"  DB       : {len(lookup)} items")
    return lookup


def load_ppi(path, label):
    try:
        df = _read_excel(path, ["HS Code"])
    except ValueError:
        df = pd.read_excel(path, header=None, dtype=str).fillna("")
        names = ["HS Code", "Designation"]
        df.columns = names[:len(df.columns)]
    result = {normalize_hs(r["HS Code"]) for _, r in df.iterrows()} - {"-"}
    print(f"  PPI_{label} : {len(result)} authorized HS codes")
    return result

# ── Check logic ───────────────────────────────────────────────────────────────

def run_hs_check(df_input, db_lookup):
    df = df_input.copy()
    for col in ("Line N°", "Part No", "Designation", "Class", "HS Code"):
        if col not in df.columns:
            df[col] = ""
    df["Part No"] = df["Part No"].fillna("").astype(str).str.strip()
    df = df[~df["Part No"].str.lower().isin(("", "nan", "none"))].reset_index(drop=True)

    df["HS Invoice"]       = df["HS Code"].apply(normalize_hs)
    df["DB Historical HS"] = df["Part No"].map(db_lookup).fillna("-")
    df["HS Verdict"]       = np.select(
        [df["DB Historical HS"] == "-", df["HS Invoice"] == df["DB Historical HS"]],
        ["NEW ITEM",                     "MATCH"],
        default="MISMATCH",
    )
    df["Volatile Flag"] = np.where(df["HS Verdict"] == "MISMATCH", "CHECK - HS MISMATCH", "-")

    cols = ["Line N°", "Part No", "Designation", "Class",
            "HS Invoice", "DB Historical HS", "HS Verdict", "Volatile Flag"]
    if "Qty" in df.columns:
        cols.insert(cols.index("Designation") + 1, "Qty")
    return df[cols].copy()


def run_ppi_check(df_input, bank, ppi_cvc, ppi_hsbc):
    ppi_set = {"CVC": ppi_cvc, "HSBC": ppi_hsbc}.get(bank.upper(), set())
    df = df_input.copy()
    for col in ("Line N°", "Part No", "Designation", "Class", "HS Code"):
        if col not in df.columns:
            df[col] = ""
    df["Part No"] = df["Part No"].fillna("").astype(str).str.strip()
    df = df[~df["Part No"].str.lower().isin(("", "nan", "none"))].reset_index(drop=True)

    df["HS Invoice"] = df["HS Code"].apply(normalize_hs)
    df["Bank"]       = bank.upper()
    df["PPI Status"] = np.select(
        [df["HS Invoice"] == "-", df["HS Invoice"].isin(ppi_set)],
        ["NO HS CODE",            "IN PPI"],
        default="NOT IN PPI",
    )

    cols = ["Line N°", "Part No", "Designation", "Class",
            "HS Invoice", "Bank", "PPI Status"]
    if "Qty" in df.columns:
        cols.insert(cols.index("Designation") + 1, "Qty")
    return df[cols].copy()


def build_tho_output(df_hs, df_ppi):
    # Both built from same df_input with identical row filter → safe positional join
    merged = pd.concat(
        [df_hs.reset_index(drop=True), df_ppi[["Bank", "PPI Status"]].reset_index(drop=True)],
        axis=1,
    )
    merged["PPI Status"] = merged["PPI Status"].fillna("—")
    merged["Bank"]       = merged["Bank"].fillna("—")

    flagged = merged[
        merged["HS Verdict"].isin(("MISMATCH", "NEW ITEM")) |
        (merged["PPI Status"] == "NOT IN PPI")
    ].copy()

    flagged["Action Requise"] = np.select(
        [flagged["HS Verdict"] == "NEW ITEM",
         flagged["HS Verdict"] == "MISMATCH",
         flagged["PPI Status"] == "NOT IN PPI"],
        ["Nouvel article — Validation requise",
         "Corriger code HS — voir suggestion DB",
         "Code HS non autorisé dans PPI"],
        default="-",
    )
    out_cols = ["Line N°", "Part No", "Designation", "Class",
                "HS Invoice", "DB Historical HS", "HS Verdict",
                "Bank", "PPI Status", "Action Requise"]
    if "Qty" in merged.columns:
        out_cols.insert(out_cols.index("Designation") + 1, "Qty")
    return flagged[out_cols]

# ── Excel output helpers ──────────────────────────────────────────────────────

def _write_df(ws, df, sheet_name):
    hfill = HEADER_FILLS[sheet_name]
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, col)
        c.fill = hfill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.alignment = Alignment(vertical="center")
            c.border = THIN
            if ri % 2 == 0:
                c.fill = GREY_FILL


def _color_col(ws, df, col_name):
    if col_name not in df.columns:
        return
    ci = list(df.columns).index(col_name) + 1
    for ri, val in enumerate(df[col_name], 2):
        style = VERDICT_STYLE.get(str(val))
        if style:
            ws.cell(ri, ci).fill, ws.cell(ri, ci).font = style


def _autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, 10), 42)


def _setup(ws):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def write_dashboard(ws, summary, invoice_name, run_time, bank):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    tc = ws["A1"]
    tc.value = "HS COMPLIANCE CHECKER — DASHBOARD"
    tc.font  = Font(bold=True, size=14, color="FFFFFF")
    tc.fill  = _fill("2E4057")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    for i, (k, v) in enumerate([
        ("Invoice",    invoice_name),
        ("Bank",       bank),
        ("Run",        run_time),
    ], 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)

    for ci, h in enumerate(["Metric", "Count", "Rate"], 1):
        c = ws.cell(7, ci, h)
        c.fill = _fill("2E4057")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN

    stats = [
        ("Total Lines",      summary["total_lines"],   ""),
        ("HS — MATCH",       summary["hs_match"],      f"{summary['hs_pass_rate']:.1%}"),
        ("HS — MISMATCH",    summary["hs_mismatch"],   ""),
        ("HS — NEW ITEM",    summary["new_items"],     ""),
        ("PPI — IN PPI",     summary["ppi_compliant"], f"{summary['ppi_pass_rate']:.1%}"),
        ("PPI — NOT IN PPI", summary["not_in_ppi"],    ""),
    ]
    fills = {
        "HS — MATCH": GREEN_FILL,  "HS — MISMATCH": RED_FILL,
        "HS — NEW ITEM": ORANGE_FILL, "PPI — IN PPI": GREEN_FILL,
        "PPI — NOT IN PPI": RED_FILL,
    }
    fonts = {
        "HS — MATCH": GREEN_FONT,  "HS — MISMATCH": RED_FONT,
        "HS — NEW ITEM": ORANGE_FONT, "PPI — IN PPI": GREEN_FONT,
        "PPI — NOT IN PPI": RED_FONT,
    }
    for ri, (label, count, rate) in enumerate(stats, 8):
        for ci, val in enumerate([label, count, rate], 1):
            c = ws.cell(ri, ci, val)
            c.border = THIN
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        if label in fills:
            for ci in (1, 2, 3):
                ws.cell(ri, ci).fill = fills[label]
                ws.cell(ri, ci).font = fonts[label]

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10


def write_hs_check(ws, df):
    _setup(ws); _write_df(ws, df, "HS_CHECK")
    _color_col(ws, df, "HS Verdict"); _color_col(ws, df, "Volatile Flag"); _autofit(ws)


def write_ppi_check(ws, df):
    _setup(ws); _write_df(ws, df, "PPI_CHECK")
    _color_col(ws, df, "PPI Status"); _autofit(ws)


def write_tho_output(ws, df):
    _setup(ws); _write_df(ws, df, "THO_OUTPUT")
    _color_col(ws, df, "HS Verdict"); _color_col(ws, df, "PPI Status"); _autofit(ws)


def write_ppi_ref(ws, ppi_path):
    _setup(ws)
    try:
        df = _read_excel(ppi_path, ["HS Code"])
    except ValueError:
        df = pd.read_excel(ppi_path, header=None, dtype=str).fillna("")
        df.columns = ["HS Code", "Designation"][:len(df.columns)]
    df = df[[c for c in df.columns if not str(c).startswith("Unnamed")]]
    if "HS Code" in df.columns:
        df["HS Code"] = df["HS Code"].apply(normalize_hs)
        df = df[df["HS Code"] != "-"].reset_index(drop=True)
    _write_df(ws, df, "PPI_REF")
    _autofit(ws)


# ── Main ──────────────────────────────────────────────────────────────────────

def abort(msg):
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk(); root.withdraw()
        messagebox.showerror("HS Checker — Error", msg)
        root.destroy()
    except Exception:
        print(f"\nERROR: {msg}")
    sys.exit(1)


def main():
    print("=" * 60)
    print("  HS COMPLIANCE CHECKER")
    print("=" * 60)

    # ── Check fixed files exist ───────────────────────────────────
    for label, path in [("DB", DB_FILE), ("PPI_CVC", PPI_CVC), ("PPI_HSBC", PPI_HSBC)]:
        if not path.exists():
            abort(f"{label} file not found: {path}\n  Place the file there and try again.")

    # ── Find latest invoice in incoming/ ─────────────────────────
    invoices = sorted(
        [f for f in INCOMING.iterdir() if f.suffix == ".xlsx" and not f.name.startswith("~$")],
        key=lambda f: f.stat().st_mtime, reverse=True,
    )
    if not invoices:
        abort(f"No invoice found in:\n  {INCOMING}\n  Drop your invoice xlsx there and try again.")

    invoice_path = invoices[0]

    print(f"\n  Invoice  : {invoice_path.name}")
    print(f"  DB       : {DB_FILE.name}")
    print(f"  PPI CVC  : {PPI_CVC.name}")
    print(f"  PPI HSBC : {PPI_HSBC.name}")

    if len(invoices) > 1:
        print(f"\n  NOTE: {len(invoices)} invoices found in incoming/ — using the most recent.")
        print(f"  Others: {[f.name for f in invoices[1:]]}")

    # ── Bank selection — clean two-button popup ───────────────────
    bank = None
    try:
        import tkinter as tk
        root = tk.Tk()
        root.title("HS Checker — Bank Selection")
        root.resizable(False, False)
        root.attributes("-topmost", True)

        # Center on screen
        root.update_idletasks()
        w, h = 340, 160
        x = (root.winfo_screenwidth()  - w) // 2
        y = (root.winfo_screenheight() - h) // 2
        root.geometry(f"{w}x{h}+{x}+{y}")

        tk.Label(
            root, text="Which bank is this invoice for?",
            font=("Segoe UI", 11, "bold"), pady=18
        ).pack()

        btn_frame = tk.Frame(root)
        btn_frame.pack()

        def pick(b):
            nonlocal bank
            bank = b
            root.destroy()

        tk.Button(
            btn_frame, text="CVC  —  Spare Parts", width=16, height=2,
            font=("Segoe UI", 10), bg="#2E75B6", fg="white",
            relief="flat", cursor="hand2",
            command=lambda: pick("CVC")
        ).pack(side="left", padx=12)

        tk.Button(
            btn_frame, text="HSBC  —  Kits", width=16, height=2,
            font=("Segoe UI", 10), bg="#70AD47", fg="white",
            relief="flat", cursor="hand2",
            command=lambda: pick("HSBC")
        ).pack(side="left", padx=12)

        root.mainloop()

        if not bank:
            abort("No bank selected.")
    except Exception:
        print("\n  Which bank is this invoice for?")
        print("  [1] CVC  (Spare Parts)")
        print("  [2] HSBC (Kits)")
        while True:
            choice = input("  Enter 1 or 2: ").strip()
            if choice == "1": bank = "CVC";  break
            if choice == "2": bank = "HSBC"; break
            print("  Please enter 1 or 2.")
    print(f"  Bank     : {bank}\n")

    # ── Load ──────────────────────────────────────────────────────
    print("  Loading...")
    try:
        df_input  = load_invoice(invoice_path)
        db_lookup = load_db(DB_FILE)
        ppi_cvc   = load_ppi(PPI_CVC,  "CVC")
        ppi_hsbc  = load_ppi(PPI_HSBC, "HSBC")
    except Exception as e:
        abort(str(e))

    # ── Run checks ────────────────────────────────────────────────
    print("\n  Running checks...")
    df_hs  = run_hs_check(df_input, db_lookup)
    df_ppi = run_ppi_check(df_input, bank, ppi_cvc, ppi_hsbc)
    df_tho = build_tho_output(df_hs, df_ppi)

    total = len(df_hs)
    summary = {
        "total_lines":   total,
        "hs_match":      int((df_hs["HS Verdict"] == "MATCH").sum()),
        "hs_mismatch":   int((df_hs["HS Verdict"] == "MISMATCH").sum()),
        "new_items":     int((df_hs["HS Verdict"] == "NEW ITEM").sum()),
        "ppi_compliant": int((df_ppi["PPI Status"] == "IN PPI").sum()),
        "not_in_ppi":    int((df_ppi["PPI Status"] == "NOT IN PPI").sum()),
        "hs_pass_rate":  round((df_hs["HS Verdict"] == "MATCH").sum() / total, 4) if total else 0,
        "ppi_pass_rate": round((df_ppi["PPI Status"] == "IN PPI").sum() / total, 4) if total else 0,
    }

    run_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    print(f"\n  {'─'*44}")
    print(f"  Total  : {summary['total_lines']} lines")
    print(f"  HS     : {summary['hs_match']} MATCH  |  {summary['hs_mismatch']} MISMATCH  |  {summary['new_items']} NEW ITEM")
    print(f"  PPI    : {summary['ppi_compliant']} IN PPI  |  {summary['not_in_ppi']} NOT IN PPI")
    print(f"  Rate   : HS {summary['hs_pass_rate']:.1%}  |  PPI {summary['ppi_pass_rate']:.1%}")
    print(f"  Flagged: {len(df_tho)} lines for review")
    print(f"  {'─'*44}")

    # ── Write output ──────────────────────────────────────────────
    month_dir = OUTPUT_DIR / datetime.now().strftime("%Y-%m")
    month_dir.mkdir(parents=True, exist_ok=True)

    stem = invoice_path.stem.replace(" ", "_")
    out_name = f"THO_CHECK_{stem}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    out_path = month_dir / out_name

    print(f"\n  Writing → output/{datetime.now().strftime('%Y-%m')}/{out_name}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, writer, df_arg in [
        ("DASHBOARD",  write_dashboard, None),
        ("HS_CHECK",   write_hs_check,  df_hs),
        ("PPI_CHECK",  write_ppi_check, df_ppi),
        ("THO_OUTPUT", write_tho_output, df_tho),
    ]:
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = TAB_COLORS[name]
        if name == "DASHBOARD":
            write_dashboard(ws, summary, invoice_path.name, run_time, bank)
        else:
            writer(ws, df_arg)

    ws = wb.create_sheet("PPI_REF")
    ws.sheet_properties.tabColor = TAB_COLORS["PPI_REF"]
    write_ppi_ref(ws, PPI_CVC if bank == "CVC" else PPI_HSBC)

    wb.save(out_path)

    print(f"\n  [DONE]  {out_path}")

    try:
        os.startfile(out_path)
        print("  Opening Excel...")
    except Exception:
        pass



if __name__ == "__main__":
    main()
