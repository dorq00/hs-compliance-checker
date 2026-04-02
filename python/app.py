"""
HS Checker — app.py  ·  Streamlit web UI
Run with:  streamlit run app.py
"""

import sys
import io
import contextlib
import traceback
from datetime import datetime
from pathlib import Path

import streamlit as st

# ── Page config (must be first Streamlit call) ────────────────────────────────
st.set_page_config(
    page_title="HS Checker",
    page_icon="⬛",
    layout="wide",
    initial_sidebar_state="expanded",
)

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

# ── Dependencies ───────────────────────────────────────────────────────────────
try:
    import pandas as pd
    import openpyxl
except ImportError as e:
    st.error(f"Missing: `{e}` — run `pip install pandas openpyxl`")
    st.stop()

try:
    from run import (
        load_invoice, load_db, load_ppi,
        run_hs_check, run_ppi_check, build_tho_output,
        write_dashboard, write_hs_check, write_ppi_check,
        write_tho_output, write_ppi_ref,
        DB_FILE, PPI_CVC, PPI_HSBC, INCOMING, OUTPUT_DIR, TAB_COLORS,
    )
except Exception as e:
    st.error(f"**Cannot import run.py:** `{e}`  \nEnsure `app.py` lives beside `run.py`.")
    st.stop()

try:
    from ppi_dashboard import (
        load_quota, load_consumption,
        build_dashboard as build_ppi_dashboard,
        write_excel as write_ppi_excel,
        APPROVAL_FILE,
    )
    DASH_AVAILABLE = True
except Exception:
    DASH_AVAILABLE = False

# ── Session state init ─────────────────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = None
if "last_error" not in st.session_state:
    st.session_state.last_error = None

# ══════════════════════════════════════════════════════════════════════════════
# THEME — Deep navy ops console, electric teal accent
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=JetBrains+Mono:ital,wght@0,300;0,400;0,500;0,700;1,400&display=swap');

:root {
  --bg0:    #06080E;
  --bg1:    #0B0F1A;
  --bg2:    #101522;
  --bg3:    #161E2E;
  --border: #1A2438;
  --brd2:   #22334E;
  --teal:   #00E8C8;
  --green:  #0FC87A;
  --red:    #FF3D5E;
  --amber:  #FFAD1F;
  --blue:   #4B8EFF;
  --purple: #9B6BFF;
  --text:   #C8D8EC;
  --text2:  #7A95B8;
  --muted:  #3E5470;
  --ui:     'Syne','Segoe UI',system-ui,sans-serif;
  --mono:   'JetBrains Mono','Cascadia Code','Consolas',monospace;
}

/* ── Base ──────────────────────────────────────────────────────────────────── */
.stApp, html, body { background: var(--bg0) !important; }
.main .block-container { padding: 0 2rem 4rem !important; max-width: 1440px !important; }
*, p, div, span, label { font-family: var(--ui) !important; }
h1,h2,h3,h4,h5,h6 { font-family: var(--ui) !important; letter-spacing: -.025em; }

/* ── Sidebar ───────────────────────────────────────────────────────────────── */
[data-testid="stSidebar"] {
  background: var(--bg1) !important;
  border-right: 1px solid var(--border) !important;
}
[data-testid="stSidebar"] > div { padding-top: 0 !important; }

/* ── Chrome ────────────────────────────────────────────────────────────────── */
#MainMenu, footer { visibility: hidden !important; }
[data-testid="stDecoration"] { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }

/* ── Primary button ────────────────────────────────────────────────────────── */
.stButton > button {
  width: 100% !important;
  background: var(--teal) !important;
  color: #060810 !important;
  border: none !important;
  border-radius: 0 !important;
  font-weight: 800 !important;
  font-size: .75rem !important;
  letter-spacing: .18em !important;
  text-transform: uppercase !important;
  font-family: var(--mono) !important;
  padding: .8rem 1.5rem !important;
  transition: all .12s ease !important;
}
.stButton > button:hover {
  background: #00FFE0 !important;
  box-shadow: 0 0 28px rgba(0,232,200,.4) !important;
}
.stButton > button:disabled {
  background: var(--border) !important;
  color: var(--muted) !important;
  box-shadow: none !important;
}

/* ── Secondary / clear button ──────────────────────────────────────────────── */
[data-testid="stBaseButton-secondary"],
button[kind="secondary"] {
  background: transparent !important;
  border: 1px solid var(--brd2) !important;
  color: var(--muted) !important;
  font-size: .65rem !important;
  padding: .35rem 1rem !important;
  letter-spacing: .12em !important;
  box-shadow: none !important;
}
[data-testid="stBaseButton-secondary"]:hover,
button[kind="secondary"]:hover {
  border-color: var(--red) !important;
  color: var(--red) !important;
  background: rgba(255,61,94,.06) !important;
  box-shadow: none !important;
}

/* ── Download button ───────────────────────────────────────────────────────── */
.stDownloadButton > button {
  width: 100% !important;
  background: transparent !important;
  border: 1px solid var(--teal) !important;
  color: var(--teal) !important;
  border-radius: 0 !important;
  font-family: var(--mono) !important;
  font-weight: 600 !important;
  font-size: .75rem !important;
  letter-spacing: .14em !important;
  text-transform: uppercase !important;
  padding: .7rem 1.5rem !important;
  transition: all .12s !important;
}
.stDownloadButton > button:hover {
  background: rgba(0,232,200,.08) !important;
  box-shadow: 0 0 20px rgba(0,232,200,.25) !important;
}

/* ── Radio ─────────────────────────────────────────────────────────────────── */
[data-testid="stRadio"] label { color: var(--text) !important; font-size: .88rem !important; }
[data-testid="stRadio"] [role="radiogroup"] { gap: .4rem !important; }

/* ── File uploader ─────────────────────────────────────────────────────────── */
[data-testid="stFileUploader"] section {
  background: var(--bg2) !important;
  border: 1px dashed var(--brd2) !important;
  border-radius: 0 !important;
  transition: border-color .15s !important;
}
[data-testid="stFileUploader"] section:hover { border-color: var(--teal) !important; }

/* ── Progress ──────────────────────────────────────────────────────────────── */
.stProgress > div { background: var(--border) !important; height: 2px !important; border-radius: 0 !important; }
.stProgress > div > div { background: linear-gradient(90deg,var(--teal),#00FFE0) !important; border-radius: 0 !important; }

/* ── Status ────────────────────────────────────────────────────────────────── */
[data-testid="stStatusContainer"] {
  background: var(--bg2) !important;
  border: 1px solid var(--border) !important;
  border-radius: 0 !important;
}

/* ── Tabs ──────────────────────────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
  background: transparent !important;
  border-bottom: 1px solid var(--border) !important;
  gap: 0 !important;
  padding: 0 !important;
}
.stTabs [data-baseweb="tab"] {
  color: var(--muted) !important;
  font-family: var(--mono) !important;
  font-size: .7rem !important;
  text-transform: uppercase !important;
  letter-spacing: .14em !important;
  padding: .75rem 1.75rem !important;
  background: transparent !important;
  border: none !important;
  border-bottom: 2px solid transparent !important;
  border-radius: 0 !important;
}
.stTabs [aria-selected="true"] { color: var(--teal) !important; border-bottom-color: var(--teal) !important; }
.stTabs [data-baseweb="tab-panel"] { padding: 1.5rem 0 !important; }

/* ── Alerts ────────────────────────────────────────────────────────────────── */
[data-baseweb="notification"] { border-radius: 0 !important; background: var(--bg2) !important; }

/* ── Expander ──────────────────────────────────────────────────────────────── */
[data-testid="stExpander"] {
  background: var(--bg2) !important;
  border: 1px solid var(--border) !important;
  border-radius: 0 !important;
}
[data-testid="stExpander"] summary { color: var(--text2) !important; }

/* ── Spinner ───────────────────────────────────────────────────────────────── */
[data-testid="stSpinner"] { color: var(--teal) !important; }

/* ── Divider ───────────────────────────────────────────────────────────────── */
hr { border-color: var(--border) !important; margin: .75rem 0 !important; }

/* ── Scrollbar ─────────────────────────────────────────────────────────────── */
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: var(--bg1); }
::-webkit-scrollbar-thumb { background: var(--brd2); }
::-webkit-scrollbar-thumb:hover { background: var(--text2); }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# HTML COMPONENTS
# ══════════════════════════════════════════════════════════════════════════════

def stat_card(label: str, value, accent: str = "var(--teal)", sub: str = "") -> str:
    sub_html = (
        f'<div style="font-family:var(--mono);font-size:.65rem;color:var(--muted);'
        f'margin-top:.45rem">{sub}</div>'
    ) if sub else ""
    return f"""
<div style="background:var(--bg2);border:1px solid var(--border);
     border-left:3px solid {accent};padding:1rem 1.2rem;min-height:88px">
  <div style="font-family:var(--mono);font-size:.58rem;color:var(--muted);
       text-transform:uppercase;letter-spacing:.16em;margin-bottom:.45rem">{label}</div>
  <div style="font-family:var(--mono);font-size:1.85rem;font-weight:700;
       color:{accent};line-height:1.1;letter-spacing:-.02em">{value}</div>
  {sub_html}
</div>"""


def section_header(title: str, badge: str = "", accent: str = "var(--teal)") -> str:
    badge_html = (
        f'<span style="background:{accent}1A;color:{accent};font-family:var(--mono);'
        f'font-size:.58rem;letter-spacing:.14em;padding:.15rem .55rem;'
        f'border:1px solid {accent}33">{badge}</span>'
    ) if badge else ""
    return f"""
<div style="display:flex;align-items:center;gap:.75rem;
     margin:1.75rem 0 1rem;border-bottom:1px solid var(--border);padding-bottom:.6rem">
  <div style="width:3px;height:1rem;background:{accent};flex-shrink:0"></div>
  <span style="font-size:.65rem;font-family:var(--mono);text-transform:uppercase;
        letter-spacing:.22em;color:var(--text2)">{title}</span>
  {badge_html}
</div>"""


def file_badge(label: str, path) -> str:
    exists = Path(path).exists() if path else False
    color  = "var(--green)" if exists else "var(--red)"
    status = "OK" if exists else "MISSING"
    return (
        f'<div style="display:flex;align-items:center;gap:.5rem;padding:.22rem 0;'
        f'font-family:var(--mono);font-size:.7rem">'
        f'<span style="color:{color};font-size:.55rem">●</span>'
        f'<span style="color:var(--text2);flex:1">{label}</span>'
        f'<span style="color:{color};letter-spacing:.06em">{status}</span>'
        f'</div>'
    )


def page_header(title: str, sub: str = "") -> str:
    sub_html = (
        f'<span style="font-family:var(--mono);font-size:.62rem;color:var(--muted);'
        f'letter-spacing:.16em;text-transform:uppercase;margin-left:.75rem">{sub}</span>'
    ) if sub else ""
    return f"""
<div style="border-bottom:1px solid var(--border);padding:1.3rem 0;margin-bottom:0">
  <span style="font-size:1.45rem;font-weight:800;color:var(--text);
        letter-spacing:-.03em">{title}</span>{sub_html}
</div>"""


# ══════════════════════════════════════════════════════════════════════════════
# DATA HELPERS
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def _load_quota_data(approval_str: str, output_str: str):
    """Cached quota load — re-runs only if the file paths change."""
    _buf = io.StringIO()
    with contextlib.redirect_stdout(_buf):
        cvc_dict, hsbc_dict = load_quota(Path(approval_str))
        consumption = load_consumption(Path(output_str))
    return cvc_dict, hsbc_dict, consumption


@st.cache_data(show_spinner=False)
def _load_db(path_str: str) -> dict:
    _buf = io.StringIO()
    with contextlib.redirect_stdout(_buf):
        return load_db(Path(path_str))


@st.cache_data(show_spinner=False)
def _load_ppi(cvc_str: str, hsbc_str: str):
    _buf = io.StringIO()
    with contextlib.redirect_stdout(_buf):
        ppi_cvc  = load_ppi(Path(cvc_str),  "CVC")
        ppi_hsbc = load_ppi(Path(hsbc_str), "HSBC")
    return ppi_cvc, ppi_hsbc


def _style_map(styler, func, subset):
    """Compatibility shim: pandas 2.1+ uses .map(), older uses .applymap()."""
    try:
        return styler.map(func, subset=subset)
    except AttributeError:
        return styler.applymap(func, subset=subset)  # pandas < 2.1


def _rerun():
    """Compatibility shim: Streamlit 1.27+ uses st.rerun(), older uses experimental."""
    try:
        st.rerun()
    except AttributeError:
        st.experimental_rerun()


def style_results_df(df: pd.DataFrame):
    """Dark-theme color coding for verdict columns."""
    _verdict_map = {
        "MATCH":               ("#091A0F", "#0FC87A"),
        "IN PPI":              ("#091A0F", "#0FC87A"),
        "MISMATCH":            ("#1C080D", "#FF3D5E"),
        "NOT IN PPI":          ("#1C080D", "#FF3D5E"),
        "NEW ITEM":            ("#1C160A", "#FFAD1F"),
        "CHECK - HS MISMATCH": ("#1C080D", "#FF3D5E"),
        "NO HS CODE":          ("#0E1422", "#4B8EFF"),
    }

    def _fmt(val):
        if val in _verdict_map:
            bg, fg = _verdict_map[val]
            return f"background-color:{bg};color:{fg};font-weight:700"
        return ""

    verdict_cols = [c for c in df.columns
                    if c in ("HS Verdict", "Volatile Flag", "PPI Status", "Action Requise")]
    style = df.style
    if verdict_cols:
        style = _style_map(style, _fmt, verdict_cols)
    return style


def build_excel_bytes(
    df_hs, df_ppi, df_tho, summary,
    invoice_name: str, run_time: str, bank: str, out_path: Path
) -> bytes:
    """Write the Excel workbook to disk and return raw bytes for download."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name in ("DASHBOARD", "HS_CHECK", "PPI_CHECK", "THO_OUTPUT"):
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = TAB_COLORS[name]
        if name == "DASHBOARD":
            write_dashboard(ws, summary, invoice_name, run_time, bank)
        elif name == "HS_CHECK":
            write_hs_check(ws, df_hs)
        elif name == "PPI_CHECK":
            write_ppi_check(ws, df_ppi)
        elif name == "THO_OUTPUT":
            write_tho_output(ws, df_tho)

    ws = wb.create_sheet("PPI_REF")
    ws.sheet_properties.tabColor = TAB_COLORS["PPI_REF"]
    write_ppi_ref(ws, PPI_CVC if bank == "CVC" else PPI_HSBC)

    wb.save(out_path)
    with open(out_path, "rb") as f:
        return f.read()


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    # ── Wordmark ───────────────────────────────────────────────────────────────
    st.markdown("""
<div style="background:var(--bg0);border-bottom:1px solid var(--border);
     padding:1.5rem 1rem 1.2rem;margin:-1rem -1rem 1.25rem">
  <div style="font-family:var(--mono);font-size:.58rem;color:var(--muted);
       letter-spacing:.28em;text-transform:uppercase;margin-bottom:.3rem">
    HS COMPLIANCE SYSTEM
  </div>
  <div style="font-size:1.5rem;font-weight:800;color:var(--teal);
       letter-spacing:-.025em;line-height:1">
    HS CHECKER
  </div>
  <div style="font-family:var(--mono);font-size:.6rem;color:var(--muted);margin-top:.3rem">
    v1.4.0  ·  HS Compliance
  </div>
</div>
""", unsafe_allow_html=True)

    # ── Navigation ─────────────────────────────────────────────────────────────
    st.markdown(
        '<div style="font-family:var(--mono);font-size:.6rem;color:var(--muted);'
        'text-transform:uppercase;letter-spacing:.2em;margin-bottom:.4rem">Navigation</div>',
        unsafe_allow_html=True,
    )
    page = st.radio("nav", ["CHECKER", "PPI DASHBOARD"], label_visibility="collapsed")

    st.markdown("<hr>", unsafe_allow_html=True)

    if page == "CHECKER":
        # ── Bank routing ───────────────────────────────────────────────────────
        st.markdown(section_header("BANK ROUTING"), unsafe_allow_html=True)
        bank_choice = st.radio(
            "bank",
            options=["CVC  —  Spare Parts", "HSBC  —  Kits"],
            label_visibility="collapsed",
        )
        bank_code = "CVC" if bank_choice.startswith("CVC") else "HSBC"

        st.markdown("<hr>", unsafe_allow_html=True)

        # ── File upload ────────────────────────────────────────────────────────
        st.markdown(section_header("INVOICE FILE"), unsafe_allow_html=True)
        uploaded = st.file_uploader(
            "invoice",
            type=["xlsx"],
            accept_multiple_files=False,
            label_visibility="collapsed",
            help="Drop your .xlsx invoice here. Previous invoices in incoming/ are replaced.",
        )

        if uploaded:
            st.markdown(
                f'<div style="font-family:var(--mono);font-size:.7rem;color:var(--green);'
                f'margin-top:.4rem;padding:.35rem .6rem;background:rgba(15,200,122,.07);'
                f'border:1px solid rgba(15,200,122,.2)">● {uploaded.name}</div>',
                unsafe_allow_html=True,
            )

        st.markdown("<hr>", unsafe_allow_html=True)

        # ── Run button ─────────────────────────────────────────────────────────
        files_ok = DB_FILE.exists() and PPI_CVC.exists() and PPI_HSBC.exists()
        run_clicked = st.button(
            "RUN CHECK  ▶",
            disabled=(uploaded is None or not files_ok),
        )
        if not files_ok:
            st.markdown(
                '<div style="font-family:var(--mono);font-size:.65rem;color:var(--red);'
                'margin-top:.4rem">⚠ Database files missing</div>',
                unsafe_allow_html=True,
            )

        # ── Last result indicator + clear ──────────────────────────────────────
        if st.session_state.results:
            r = st.session_state.results
            st.markdown(
                f'<div style="font-family:var(--mono);font-size:.65rem;color:var(--green);'
                f'margin-top:.6rem;padding:.35rem .6rem;background:rgba(15,200,122,.07);'
                f'border:1px solid rgba(15,200,122,.18)">'
                f'✓ {r["invoice_name"]}<br>'
                f'<span style="color:var(--muted)">{r["run_time"]}  ·  {r["bank"]}</span></div>',
                unsafe_allow_html=True,
            )
            if st.button("✕  CLEAR RESULTS", key="clear", type="secondary"):
                st.session_state.results = None
                _rerun()

        st.markdown("<hr>", unsafe_allow_html=True)

        # ── DB file status ─────────────────────────────────────────────────────
        st.markdown(section_header("DATABASE FILES"), unsafe_allow_html=True)
        st.markdown(
            file_badge("db.xlsx",       DB_FILE) +
            file_badge("ppi_cvc.xlsx",  PPI_CVC) +
            file_badge("ppi_hsbc.xlsx", PPI_HSBC),
            unsafe_allow_html=True,
        )

    else:
        # PPI Dashboard sidebar
        run_clicked = False
        uploaded    = None
        bank_code   = "CVC"

        approval_path = APPROVAL_FILE if DASH_AVAILABLE else ""
        st.markdown(section_header("APPROVAL FILE"), unsafe_allow_html=True)
        st.markdown(file_badge("2026 PPI Approval…", approval_path), unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown(section_header("OUTPUT FILES"), unsafe_allow_html=True)
        n_outputs = len(list(OUTPUT_DIR.rglob("THO_CHECK_*.xlsx"))) if OUTPUT_DIR.exists() else 0
        st.markdown(
            f'<div style="font-family:var(--mono);font-size:.72rem;color:var(--text2);'
            f'padding:.25rem 0">'
            f'<span style="color:var(--teal)">{n_outputs}</span> check file(s) found in output/</div>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
# PAGE — CHECKER
# ══════════════════════════════════════════════════════════════════════════════

if page == "CHECKER":
    st.markdown(page_header("HS COMPLIANCE CHECK", "Invoice Validator"), unsafe_allow_html=True)

    # ── Run & process ──────────────────────────────────────────────────────────
    if run_clicked and uploaded:
        progress = st.progress(0)

        with st.status("Running compliance check…", expanded=True) as status:
            try:
                st.write("Saving invoice to incoming/…")
                INCOMING.mkdir(parents=True, exist_ok=True)
                invoice_path = INCOMING / uploaded.name
                with open(invoice_path, "wb") as f:
                    f.write(uploaded.getbuffer())
                progress.progress(10)

                st.write("Reading invoice…")
                _buf = io.StringIO()
                with contextlib.redirect_stdout(_buf):
                    df_input = load_invoice(invoice_path)
                progress.progress(22)

                st.write("Loading HS code database…")
                db_lookup = _load_db(str(DB_FILE))
                progress.progress(40)

                st.write("Loading PPI authorization lists…")
                ppi_cvc, ppi_hsbc = _load_ppi(str(PPI_CVC), str(PPI_HSBC))
                progress.progress(55)

                st.write("Running HS code check…")
                df_hs = run_hs_check(df_input, db_lookup)
                progress.progress(68)

                st.write("Running PPI authorization check…")
                df_ppi = run_ppi_check(df_input, bank_code, ppi_cvc, ppi_hsbc)
                progress.progress(80)

                st.write("Building output…")
                df_tho = build_tho_output(df_hs, df_ppi)
                progress.progress(90)

                total = len(df_hs)
                summary = {
                    "total_lines":   total,
                    "hs_match":      int((df_hs["HS Verdict"] == "MATCH").sum()),
                    "hs_mismatch":   int((df_hs["HS Verdict"] == "MISMATCH").sum()),
                    "new_items":     int((df_hs["HS Verdict"] == "NEW ITEM").sum()),
                    "ppi_compliant": int((df_ppi["PPI Status"] == "IN PPI").sum()),
                    "not_in_ppi":    int((df_ppi["PPI Status"] == "NOT IN PPI").sum()),
                    "hs_pass_rate":  round((df_hs["HS Verdict"] == "MATCH").sum() / total, 4) if total else 0,
                    "ppi_pass_rate": round((df_ppi["PPI Status"] == "IN PPI").sum() / total, 4) if total else 0,
                }
                run_time = datetime.now().strftime("%Y-%m-%d %H:%M")

                st.write("Writing Excel report…")
                month_dir = OUTPUT_DIR / datetime.now().strftime("%Y-%m")
                month_dir.mkdir(parents=True, exist_ok=True)
                stem     = invoice_path.stem.replace(" ", "_")
                out_name = f"THO_CHECK_{stem}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                out_path = month_dir / out_name

                excel_bytes = build_excel_bytes(
                    df_hs, df_ppi, df_tho, summary,
                    invoice_path.name, run_time, bank_code, out_path,
                )
                progress.progress(100)

                flagged = len(df_tho)
                st.session_state.results = {
                    "summary":      summary,
                    "df_hs":        df_hs,
                    "df_ppi":       df_ppi,
                    "df_tho":       df_tho,
                    "excel_bytes":  excel_bytes,
                    "out_name":     out_name,
                    "invoice_name": uploaded.name,
                    "bank":         bank_code,
                    "run_time":     run_time,
                    "flagged":      flagged,
                }

                st.session_state.last_error = None
                label = f"✓ Check complete — {flagged} item{'s' if flagged != 1 else ''} flagged for review"
                status.update(label=label, state="complete", expanded=False)

            except Exception as e:
                progress.empty()
                status.update(label="✗ Check failed", state="error", expanded=True)
                st.session_state.last_error = (str(e), traceback.format_exc())
                st.session_state.results = None

    # ── Render results ─────────────────────────────────────────────────────────
    results = st.session_state.get("results")

    if st.session_state.last_error and not results:
        err_msg, err_tb = st.session_state.last_error
        st.error(f"**Last run failed:** {err_msg}")
        with st.expander("Traceback"):
            st.code(err_tb)

    if not results:
        st.markdown("""
<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
     padding:5rem 2rem;text-align:center;user-select:none">
  <div style="width:52px;height:52px;border:1px solid var(--brd2);
       display:flex;align-items:center;justify-content:center;margin-bottom:1.75rem;
       background:var(--bg2)">
    <span style="font-size:1.4rem;opacity:.25">⬛</span>
  </div>
  <div style="font-family:var(--mono);font-size:.7rem;color:var(--muted);
       letter-spacing:.2em;text-transform:uppercase;line-height:2.2">
    Upload an invoice<br>Select a bank<br>Click RUN CHECK
  </div>
</div>
""", unsafe_allow_html=True)

    else:
        summary      = results["summary"]
        df_hs        = results["df_hs"]
        df_ppi       = results["df_ppi"]
        df_tho       = results["df_tho"]
        excel_bytes  = results["excel_bytes"]
        out_name     = results["out_name"]
        invoice_name = results["invoice_name"]
        bank         = results["bank"]
        run_time     = results["run_time"]
        flagged      = results["flagged"]

        # ── Summary cards ──────────────────────────────────────────────────────
        st.markdown(
            section_header("SUMMARY", f"{summary['total_lines']} LINES  ·  {bank}"),
            unsafe_allow_html=True,
        )

        c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
        with c1: st.markdown(stat_card("TOTAL LINES",    summary["total_lines"],                "var(--text2)"),  unsafe_allow_html=True)
        with c2: st.markdown(stat_card("HS MATCH",       f"{summary['hs_pass_rate']:.1%}",      "var(--green)",  f"{summary['hs_match']} lines"), unsafe_allow_html=True)
        with c3: st.markdown(stat_card("HS MISMATCH",    summary["hs_mismatch"],                "var(--red)"),    unsafe_allow_html=True)
        with c4: st.markdown(stat_card("NEW ITEMS",      summary["new_items"],                  "var(--amber)",  "not in DB"), unsafe_allow_html=True)
        with c5: st.markdown(stat_card("PPI AUTH",       f"{summary['ppi_pass_rate']:.1%}",     "var(--green)",  f"{summary['ppi_compliant']} lines"), unsafe_allow_html=True)
        with c6: st.markdown(stat_card("NOT IN PPI",     summary["not_in_ppi"],                 "var(--red)"),    unsafe_allow_html=True)
        with c7: st.markdown(stat_card("FLAGGED",        flagged,                               "var(--purple)", "for review"), unsafe_allow_html=True)

        # ── Export row ─────────────────────────────────────────────────────────
        st.markdown(section_header("EXPORT"), unsafe_allow_html=True)

        meta_col, dl_col = st.columns([3, 1])
        with meta_col:
            st.markdown(f"""
<div style="background:var(--bg2);border:1px solid var(--border);padding:.8rem 1.25rem;
     display:flex;gap:2.5rem;align-items:center;flex-wrap:wrap">
  <div>
    <div style="font-family:var(--mono);font-size:.55rem;color:var(--muted);
         letter-spacing:.14em;text-transform:uppercase;margin-bottom:.2rem">Invoice</div>
    <div style="font-family:var(--mono);font-size:.8rem;color:var(--text)">{invoice_name}</div>
  </div>
  <div>
    <div style="font-family:var(--mono);font-size:.55rem;color:var(--muted);
         letter-spacing:.14em;text-transform:uppercase;margin-bottom:.2rem">Bank</div>
    <div style="font-family:var(--mono);font-size:.8rem;color:var(--teal);font-weight:700">{bank}</div>
  </div>
  <div>
    <div style="font-family:var(--mono);font-size:.55rem;color:var(--muted);
         letter-spacing:.14em;text-transform:uppercase;margin-bottom:.2rem">Run Time</div>
    <div style="font-family:var(--mono);font-size:.8rem;color:var(--text)">{run_time}</div>
  </div>
  <div>
    <div style="font-family:var(--mono);font-size:.55rem;color:var(--muted);
         letter-spacing:.14em;text-transform:uppercase;margin-bottom:.2rem">Output File</div>
    <div style="font-family:var(--mono);font-size:.7rem;color:var(--text2)">{out_name}</div>
  </div>
</div>
""", unsafe_allow_html=True)
        with dl_col:
            st.markdown('<div style="height:.8rem"></div>', unsafe_allow_html=True)
            st.download_button(
                "⬇  DOWNLOAD EXCEL",
                data=excel_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # ── Result tabs ────────────────────────────────────────────────────────
        st.markdown(section_header("RESULTS"), unsafe_allow_html=True)

        mismatch_label = f"{summary['hs_mismatch']} MISMATCH · {summary['new_items']} NEW"
        tab_hs, tab_ppi, tab_tho = st.tabs([
            f"HS CHECK  ({mismatch_label})",
            f"PPI CHECK  ({summary['not_in_ppi']} NOT IN PPI)",
            f"FLAGGED OUTPUT  ({flagged} FLAGGED)",
        ])

        with tab_hs:
            st.dataframe(style_results_df(df_hs), use_container_width=True, height=460)

        with tab_ppi:
            st.dataframe(style_results_df(df_ppi), use_container_width=True, height=460)

        with tab_tho:
            if df_tho.empty:
                st.markdown("""
<div style="text-align:center;padding:2.5rem;font-family:var(--mono);font-size:.78rem;
     color:var(--green);letter-spacing:.1em">✓  NO ITEMS FLAGGED — ALL LINES CLEAR</div>
""", unsafe_allow_html=True)
            else:
                st.dataframe(style_results_df(df_tho), use_container_width=True, height=460)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE — PPI DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

elif page == "PPI DASHBOARD":
    st.markdown(page_header("PPI QUOTA DASHBOARD", "Q1 2026"), unsafe_allow_html=True)

    if not DASH_AVAILABLE:
        st.error("`ppi_dashboard.py` could not be imported. Ensure it is in the same folder as `app.py`.")
        st.stop()

    if not APPROVAL_FILE.exists():
        st.warning(f"Approval file not found:  \n`{APPROVAL_FILE}`")
        st.info("Place **ppi_approval_q1_2026.xlsx** in the `ppi/` folder.")
        st.stop()

    with st.spinner("Loading quota data…"):
        try:
            cvc_dict, hsbc_dict, consumption = _load_quota_data(
                str(APPROVAL_FILE), str(OUTPUT_DIR)
            )
            df_quota = build_ppi_dashboard(cvc_dict, hsbc_dict, consumption)
        except SystemExit:
            st.error("Failed to load approval file. Check sheet name and column structure.")
            st.stop()
        except Exception as e:
            st.error(f"Error loading dashboard: {e}")
            st.stop()

    # ── Quota summary cards ────────────────────────────────────────────────────
    total_cvc_auth  = int(df_quota["Q1 Auth CVC"].sum())
    total_hsbc_auth = int(df_quota["Q1 Auth HSBC"].sum())
    consumed_cvc    = int(df_quota["Consumed CVC"].sum())
    consumed_hsbc   = int(df_quota["Consumed HSBC"].sum())
    n_cvc_hs        = int((df_quota["Q1 Auth CVC"]  > 0).sum())
    n_hsbc_hs       = int((df_quota["Q1 Auth HSBC"] > 0).sum())
    pct_cvc         = consumed_cvc  / total_cvc_auth  * 100 if total_cvc_auth  else 0
    pct_hsbc        = consumed_hsbc / total_hsbc_auth * 100 if total_hsbc_auth else 0

    def _quota_accent(pct):
        if pct >= 90: return "var(--red)"
        if pct >= 70: return "var(--amber)"
        return "var(--green)"

    st.markdown(section_header("QUOTA SUMMARY", "Q1 2026", "var(--purple)"), unsafe_allow_html=True)

    q1, q2, q3, q4, q5, q6 = st.columns(6)
    with q1: st.markdown(stat_card("CVC HS CODES",    n_cvc_hs,       "var(--blue)"),                                   unsafe_allow_html=True)
    with q2: st.markdown(stat_card("CVC AUTH QTY",    total_cvc_auth, "var(--blue)"),                                   unsafe_allow_html=True)
    with q3: st.markdown(stat_card("CVC CONSUMED",    consumed_cvc,   _quota_accent(pct_cvc),  f"{pct_cvc:.1f}% used"), unsafe_allow_html=True)
    with q4: st.markdown(stat_card("HSBC HS CODES",   n_hsbc_hs,      "var(--purple)"),                                 unsafe_allow_html=True)
    with q5: st.markdown(stat_card("HSBC AUTH QTY",   total_hsbc_auth,"var(--purple)"),                                 unsafe_allow_html=True)
    with q6: st.markdown(stat_card("HSBC CONSUMED",   consumed_hsbc,  _quota_accent(pct_hsbc), f"{pct_hsbc:.1f}% used"), unsafe_allow_html=True)

    # ── Quota detail table ─────────────────────────────────────────────────────
    st.markdown(section_header("QUOTA DETAIL", f"{len(df_quota)} HS CODES", "var(--blue)"), unsafe_allow_html=True)

    def _style_quota_df(df: pd.DataFrame):
        def _pct_fmt(val):
            if not isinstance(val, (int, float)):
                return ""
            if val >= 90:
                return "color:#FF3D5E;font-weight:700;font-family:var(--mono)"
            if val >= 70:
                return "color:#FFAD1F;font-weight:600;font-family:var(--mono)"
            return "color:#0FC87A;font-family:var(--mono)"
        return _style_map(df.style, _pct_fmt, ["% CVC Used", "% HSBC Used"])

    st.dataframe(_style_quota_df(df_quota), use_container_width=True, height=520)

    # ── Export ─────────────────────────────────────────────────────────────────
    st.markdown(section_header("EXPORT"), unsafe_allow_html=True)
    _, dl_col, _ = st.columns([2, 1, 2])
    with dl_col:
        try:
            run_time_dash = datetime.now().strftime("%Y-%m-%d %H:%M")
            out_name_dash = f"PPI_QUOTA_DASHBOARD_{datetime.now().strftime('%Y-%m')}.xlsx"
            out_path_dash = OUTPUT_DIR / out_name_dash
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

            _buf2 = io.StringIO()
            with contextlib.redirect_stdout(_buf2):
                write_ppi_excel(df_quota, out_path_dash, run_time_dash)

            with open(out_path_dash, "rb") as f:
                quota_bytes = f.read()

            st.download_button(
                "⬇  DOWNLOAD QUOTA REPORT",
                data=quota_bytes,
                file_name=out_name_dash,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Export failed: {e}")
